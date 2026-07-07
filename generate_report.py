import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ========== 数据 ==========
# 近一周数据（2026-06-30 ~ 2026-07-06，不含今天7月7日的不完整数据）
dates = ['2026-06-30', '2026-07-01', '2026-07-02', '2026-07-03', '2026-07-04', '2026-07-05', '2026-07-06']
active_dau = [1240069, 1257607, 1278013, 1237187, 834422, 651250, 1196419]
community_uv = [3928, 4248, 4585, 3918, 1985, 1907, 3777]

# ========== 数据文件 ==========
save_dir = os.path.dirname(os.path.abspath(__file__))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '主动DAU与社区首页UV'

# 表头样式
header_font = Font(bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# 写表头
headers = ['日期', '主动DAU', '社区首页UV']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 写数据
for i, d in enumerate(dates):
    row = i + 2
    ws.cell(row=row, column=1, value=d).border = thin_border
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=2, value=active_dau[i]).border = thin_border
    ws.cell(row=row, column=3, value=community_uv[i]).border = thin_border

# 设置列宽
ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 14

# 周均汇总行
ws.cell(row=len(dates)+2, column=1, value='周均').font = Font(bold=True)
ws.cell(row=len(dates)+2, column=2, value=round(sum(active_dau)/7)).font = Font(bold=True)
ws.cell(row=len(dates)+2, column=3, value=round(sum(community_uv)/7)).font = Font(bold=True)

xlsx_path = os.path.join(save_dir, '主动DAU与社区首页UV数据.xlsx')
wb.save(xlsx_path)
print(f'数据已保存: {xlsx_path}')

# ========== ECharts HTML ===========
html_content = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>主动DAU & 社区首页UV - 折线图</title>
    <script src="https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js"></script>
    <style>
        body { margin: 0; padding: 20px; background: #f5f7fa; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; }
        .container { max-width: 1000px; margin: 0 auto; background: #fff; border-radius: 12px; box-shadow: 0 2px 12px rgba(0,0,0,0.0); padding: 30px; }
        h2 { text-align: center; color: #333; margin: 0 0 20px 0; font-weight: 500; }
        #chart { width: 100%; height: 500px; }
        .info { text-align: center; color: #999; font-size: 12px; margin-top: 15px; }
    </style>
</head>
<body>
    <div class="container">
        <h2>主动DAU & 社区首页UV - 近一周趋势</h2>
        <div id="chart"></div>
        <div class="info">数据周期: 2026-06-30 ~ 2026-07-06</div>
    </div>
    <script>
        var chartDom = document.getElementById('chart');
        var myChart = echarts.init(chartDom);
        var option = {
            tooltip: {
                trigger: 'axis',
                axisPointer: { type: 'cross' },
                backgroundColor: 'rgba(255,255,255,0.95)',
                borderColor: '#e0e0e0',
                borderWidth: 1,
                padding: [10, 14],
                formatter: function(params) {
                    var html = '<div style="font-weight:600;margin-bottom:6px;">' + params[0].axisValue + '</div>';
                    params.forEach(function(p) {
                        var val = p.seriesIndex === 0 ? p.value.toLocaleString() : p.value;
                        html += '<div style="display:flex;align-items:center;gap:6px;margin:3px 0;">' +
                            '<span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:' + p.color + ';"></span>' +
                            p.seriesName + ': <strong>' + val + '</strong></div>';
                    });
                    return html;
                }
            },
            legend: {
                data: ['主动DAU', '社区首页UV'],
                top: 10,
                textStyle: { fontSize: 13 }
            },
            grid: {
                left: 60,
                right: 60,
                bottom: 40,
                top: 60
            },
            xAxis: {
                type: 'category',
                data: ''' + str(dates) + '''.map(d => d.slice(5)),
                axisLabel: {
                    fontSize: 12,
                    fontWeight: 'bold'
                },
                axisLine: { lineStyle: { color: '#ccc' } }
            },
            yAxis: [
                {
                    type: 'value',
                    name: '主动DAU',
                    nameTextStyle: { fontSize: 13, fontWeight: 'bold', padding: [0, 0, 0, 50] },
                    min: 0,
                    max: 1600000,
                    splitLine: { lineStyle: { color: '#f0f0f0', type: 'dashed' } },
                    axisLabel: {
                        fontSize: 11,
                        formatter: function(value) {
                            return value >= 10000 ? (value / 10000).toFixed(0) + '万' : value;
                        }
                    }
                },
                {
                    type: 'value',
                    name: '社区首页UV',
                    nameTextStyle: { fontSize: 13, fontWeight: 'bold', padding: [0, 50, 0, 0] },
                    min: 0,
                    max: 6000,
                    splitLine: { show: false },
                    axisLabel: { fontSize: 11 }
                }
            ],
            series: [
                {
                    name: '主动DAU',
                    type: 'line',
                    yAxisIndex: 0,
                    data: ''' + str(active_dau) + ''',
                    smooth: true,
                    symbol: 'circle',
                    symbolSize: 8,
                    lineStyle: { width: 3, color: '#5470C6' },
                    itemStyle: { color: '#5470C6' },
                    areaStyle: {
                        color: {
                            type: 'linear',
                            x: 0, y: 0, x2: 0, y2: 1,
                            colorStops: [
                                { offset: 0, color: 'rgba(84,112,198,0.25)' },
                                { offset: 1, color: 'rgba(84,112,198,0.02)' }
                            ]
                        }
                    },
                    markLine: {
                        silent: true,
                        data: [{ type: 'average', name: '周均' }],
                        label: {
                            formatter: '周均: ' + Math.round(''' + str(sum(active_dau)/7) + ''').toLocaleString(),
                            fontSize: 11, color: '#5470C6'
                        },
                        lineStyle: { color: '#5470C6', type: 'dashed', width: 1.5 }
                    }
                },
                {
                    name: '社区首页UV',
                    type: 'line',
                    yAxisIndex: 1,
                    data: ''' + str(community_uv) + ''',
                    smooth: true,
                    symbol: 'diamond',
                    symbolSize: 8,
                    lineStyle: { width: 3, color: '#EE6666' },
                    itemStyle: { color: '#EE6666' },
                    areaStyle: {
                        color: {
                            type: 'linear',
                            x: 0, y: 0, x2: 0, y2: 1,
                            colorStops: [
                                { offset: 0, color: 'rgba(238,102,102,0.25)' },
                                { offset: 1, color: 'rgba(238,102,102,0.02)' }
                            ]
                        }
                    },
                    markLine: {
                        silent: true,
                        data: [{ type: 'average', name: '周均' }],
                        label: {
                            formatter: '周均: ' + Math.round(''' + str(sum(community_uv)/7) + '''),
                            fontSize: 11, color: '#EE6666'
                        },
                        lineStyle: { color: '#EE6666', type: 'dashed', width: 1.5 }
                    }
                }
            ]
        };
        myChart.setOption(option);
        window.addEventListener('resize', function() { myChart.resize(); });
    </script>
</body>
</html>'''

html_path = os.path.join(save_dir, '主动DAU与社区首页UV_折线图.html')
with open(html_path, 'w', encoding='utf-8') as f:
    f.write(html_content)
print(f'HTML已生成: {html_path}')