# -*- coding: utf-8 -*-
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ========== 原始数据（6月1日 ~ 7月7日） ==========
raw_data = [
    ("2026-06-01", 1182600, 3801),
    ("2026-06-02", 1213861, 3961),
    ("2026-06-03", 1213703, 4152),
    ("2026-06-04", 1212854, 4168),
    ("2026-06-05", 1211218, 4334),
    ("2026-06-06", 965831,  3009),
    ("2026-06-07", 722395,  2354),
    ("2026-06-08", 1261710, 4822),
    ("2026-06-09", 1257023, 4419),
    ("2026-06-10", 1310748, 5763),
    ("2026-06-11", 1320599, 5317),
    ("2026-06-12", 1321505, 5104),
    ("2026-06-13", 986432,  2705),
    ("2026-06-14", 743602,  2309),
    ("2026-06-15", 1266008, 4509),
    ("2026-06-16", 1267815, 4527),
    ("2026-06-17", 1255257, 4303),
    ("2026-06-18", 1259098, 4537),
    ("2026-06-19", 1025799, 3262),
    ("2026-06-20", 829099,  2309),
    ("2026-06-21", 760663,  2373),
    ("2026-06-22", 1235126, 4103),
    ("2026-06-23", 1276774, 4475),
    ("2026-06-24", 1290074, 4518),
    ("2026-06-25", 1289016, 4801),
    ("2026-06-26", 1266258, 4397),
    ("2026-06-27", 961133,  2555),
    ("2026-06-28", 744261,  2356),
    ("2026-06-29", 1216927, 4112),
    ("2026-06-30", 1240069, 3928),
    ("2026-07-01", 1257607, 4248),
    ("2026-07-02", 1278013, 4585),
    ("2026-07-03", 1237187, 3918),
    ("2026-07-04", 834422,  1985),
    ("2026-07-05", 651250,  1907),
    ("2026-07-06", 1196419, 3777),
]

dates_6_7 = [d[0] for d in raw_data]
dau_list = [d[1] for d in raw_data]
community_list = [d[2] for d in raw_data]

# 周均计算（每7天一周，不足7天丢弃）
def calc_weekly(data, labels):
    weeks = []
    week_labels = []
    for i in range(0, len(data) - 6, 7):
        week = data[i:i+7]
        weeks.append(round(sum(week) / 7, 1))
        week_labels.append(labels[i][:10])
    return weeks, week_labels

weekly_dau, week_labels = calc_weekly(dau_list, dates_6_7)
weekly_community, _ = calc_weekly(community_list, dates_6_7)

# 月均
monthly_data = {}
for d, dau, cu in zip(dates_6_7, dau_list, community_list):
    m = d[:7]
    if m not in monthly_data:
        monthly_data[m] = {'dau': [], 'community': []}
    monthly_data[m]['dau'].append(dau)
    monthly_data[m]['community'].append(cu)

month_labels = sorted(monthly_data.keys())
monthly_dau = [round(sum(monthly_data[m]['dau']) / len(monthly_data[m]['dau']), 1) for m in month_labels]
monthly_community = [round(sum(monthly_data[m]['community']) / len(monthly_data[m]['community']), 1) for m in month_labels]

# ========== 保存Excel ==========
save_dir = os.path.dirname(os.path.abspath(__file__))

wb = openpyxl.Workbook()
# Sheet1: 日数据
ws1 = wb.active
ws1.title = '日数据'
header_font = Font(bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col, h in enumerate(['日期', '主动DAU', '社区首页UV'], 1):
    c = ws1.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border

for i, (d, dau, cu) in enumerate(raw_data):
    ws1.cell(row=i+2, column=1, value=d).border = thin_border
    ws1.cell(row=i+2, column=1).alignment = center_align
    ws1.cell(row=i+2, column=2, value=dau).border = thin_border
    ws1.cell(row=i+2, column=3, value=cu).border = thin_border

ws1.column_dimensions['A'].width = 14
ws1.column_dimensions['B'].width = 14
ws1.column_dimensions['C'].width = 14

# Sheet2: 周均
ws2 = wb.create_sheet('周均')
for col, h in enumerate(['周期', '主动DAU(周均)', '社区首页UV(周均)'], 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border

for i, (wl, wd, wc) in enumerate(zip(week_labels, weekly_dau, weekly_community)):
    ws2.cell(row=i+2, column=1, value=wl).border = thin_border
    ws2.cell(row=i+2, column=1).alignment = center_align
    ws2.cell(row=i+2, column=2, value=wd).border = thin_border
    ws2.cell(row=i+2, column=3, value=wc).border = thin_border

ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 16
ws2.column_dimensions['C'].width = 18

# Sheet3: 月均
ws3 = wb.create_sheet('月均')
for col, h in enumerate(['月份', '主动DAU(月均)', '社区首页UV(月均)'], 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border

for i, (ml, md, mc) in enumerate(zip(month_labels, monthly_dau, monthly_community)):
    ws3.cell(row=i+2, column=1, value=ml).border = thin_border
    ws3.cell(row=i+2, column=1).alignment = center_align
    ws3.cell(row=i+2, column=2, value=md).border = thin_border
    ws3.cell(row=i+2, column=3, value=mc).border = thin_border

ws3.column_dimensions['A'].width = 10
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 18

xlsx_path = os.path.join(save_dir, '主动DAU与社区首页UV数据.xlsx')
wb.save(xlsx_path)
print(f'Excel已保存: {xlsx_path}')

# ========== 生成ECharts HTML ==========
# 日数据
days_data = {dates_6_7[i]: {'dau': dau_list[i], 'community': community_list[i]} for i in range(len(dates_6_7))}
# 周数据：按自然周聚合，格式 251201~251207
week_map = {}
for i in range(0, len(dates_6_7) - 6, 7):
    week_dates = dates_6_7[i:i+7]
    year_week = week_dates[0][2:4] + week_dates[0][5:7].replace('0','') + week_dates[0][8:10].replace('0','') + '~' + week_dates[6][5:7].replace('0','') + week_dates[6][8:10].replace('0','')
    # Recalculate weekly
    week_avg_dau = round(sum(dau_list[i:i+7]) / 7)
    week_avg_community = round(sum(community_list[i:i+7]) / 7)
    week_map[year_week] = {'dau': week_avg_dau, 'community': week_avg_community}

# 月数据
month_map = {}
for m in month_labels:
    d_list = [dau_list[j] for j in range(len(dates_6_7)) if dates_6_7[j].startswith(m)]
    c_list = [community_list[j] for j in range(len(dates_6_7)) if dates_6_7[j].startswith(m)]
    month_map[m] = {'dau': round(sum(d_list)/len(d_list)), 'community': round(sum(c_list)/len(c_list))}

import json
html = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>主动DAU & 社区首页UV</title>
<script src="https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js"></script>
<style>
* { margin: 0; padding: 0; box-sizing: border-box; }
body { background: #f5f7fa; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; padding: 20px; }
.container { max-width: 1100px; margin: 0 auto; background: #fff; border-radius: 12px; padding: 30px; }
h2 { text-align: center; color: #333; font-weight: 500; margin-bottom: 20px; }
.tabs { display: flex; justify-content: center; gap: 10px; margin-bottom: 20px; }
.tab-btn { padding: 8px 24px; border: 1px solid #d0d5dd; border-radius: 6px; background: #fff; cursor: pointer; font-size: 14px; color: #555; transition: all .2s; }
.tab-btn:hover { border-color: #5470C6; color: #5470C6; }
.tab-btn.active { background: #5470C6; color: #fff; border-color: #5470C6; }
#chart { width: 100%; height: 500px; }
.table-wrap { margin-top: 20px; overflow-x: auto; }
table { width: 100%; border-collapse: collapse; font-size: 13px; }
th { background: #f0f4ff; color: #333; font-weight: 600; padding: 10px 12px; border: 1px solid #e0e5ee; text-align: center; }
td { padding: 8px 12px; border: 1px solid #e0e5ee; text-align: center; color: #444; }
tr:nth-child(even) { background: #fafbfc; }
tr:hover { background: #f0f4ff; }
.info { text-align: center; color: #999; font-size: 12px; margin-top: 12px; }
.hidden { display: none; }
</style>
</head>
<body>
<div class="container">
  <h2>主动DAU & 社区首页UV</h2>
  <div class="tabs">
    <button class="tab-btn active" data-tab="daily">日数据</button>
    <button class="tab-btn" data-tab="weekly">周均</button>
    <button class="tab-btn" data-tab="monthly">月均</button>
  </div>
  <div id="chart"></div>
  <div class="table-wrap">
    <table id="data-table">
      <thead><tr><th>日期</th><th>主动DAU</th><th>社区首页UV</th></tr></thead>
      <tbody></tbody>
    </table>
  </div>
  <div class="info">数据周期: 2026-06-01 ~ 2026-07-06</div>
</div>
<script>
// ========== 数据 ==========
var daysData = ''' + json.dumps(days_data) + ''';
var weeksData = ''' + json.dumps(week_map) + ''';
var monthsData = ''' + json.dumps(month_map) + ''';

function buildChart(data, type) {
  var labels = Object.keys(data);
  var dauVals = labels.map(function(k) { return data[k].dau; });
  var comVals = labels.map(function(k) { return data[k].community; });

  // 手动计算均值，确保准确
  var dauSum = dauVals.reduce(function(a, b) { return a + b; }, 0);
  var comSum = comVals.reduce(function(a, b) { return a + b; }, 0);
  var dauAvg = Math.round(dauSum / dauVals.length);
  var comAvg = Math.round(comSum / comVals.length);

  var option = {
    tooltip: {
      trigger: 'axis',
      axisPointer: { type: 'cross' },
      backgroundColor: 'rgba(255,255,255,0.95)',
      borderColor: '#e0e0e0', borderWidth: 1, padding: [10, 14],
      formatter: function(params) {
        var s = '<div style="font-weight:600;margin-bottom:6px;">' + params[0].axisValue + '</div>';
        params.forEach(function(p) {
          var v = p.seriesIndex === 0 ? Number(p.value).toLocaleString() : p.value;
          s += '<div style="display:flex;align-items:center;gap:6px;margin:3px 0;">' +
            '<span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:' + p.color + ';"></span>' +
            p.seriesName + ': <strong>' + v + '</strong></div>';
        });
        return s;
      }
    },
    legend: { data: ['主动DAU', '社区首页UV'], top: 5, textStyle: { fontSize: 13 } },
    grid: { left: 60, right: 60, bottom: 30, top: 50 },
    xAxis: {
      type: 'category', data: labels,
      axisLabel: { fontSize: 11, rotate: type === 'daily' ? 45 : 0 },
      axisLine: { lineStyle: { color: '#ccc' } }
    },
    yAxis: [
      {
        type: 'value', name: '主动DAU',
        nameTextStyle: { fontSize: 13, fontWeight: 'bold', padding: [0,0,0,40] },
        min: 0,
        splitLine: { lineStyle: { color: '#f0f0f0', type: 'dashed' } },
        axisLabel: { fontSize: 11, formatter: function(v) { return v >= 10000 ? (v/10000).toFixed(0) + '万' : v; } }
      },
      {
        type: 'value', name: '社区首页UV',
        nameTextStyle: { fontSize: 13, fontWeight: 'bold', padding: [0,40,0,0] },
        min: 0,
        splitLine: { show: false },
        axisLabel: { fontSize: 11 }
      }
    ],
    series: [
      {
        name: '主动DAU', type: 'line', yAxisIndex: 0,
        data: dauVals, smooth: true,
        symbol: 'circle', symbolSize: 6,
        lineStyle: { width: 3, color: '#5470C6' },
        itemStyle: { color: '#5470C6' },
        areaStyle: {
          color: { type: 'linear', x: 0, y: 0, x2: 0, y2: 1,
            colorStops: [
              { offset: 0, color: 'rgba(84,112,198,0.25)' },
              { offset: 1, color: 'rgba(84,112,198,0.02)' }
            ]
          }
        },
        // 不显示均值线
      },
      {
        name: '社区首页UV', type: 'line', yAxisIndex: 1,
        data: comVals, smooth: true,
        symbol: 'diamond', symbolSize: 6,
        lineStyle: { width: 3, color: '#EE6666' },
        itemStyle: { color: '#EE6666' },
        areaStyle: {
          color: { type: 'linear', x: 0, y: 0, x2: 0, y2: 1,
            colorStops: [
              { offset: 0, color: 'rgba(238,102,102,0.25)' },
              { offset: 1, color: 'rgba(238,102,102,0.02)' }
            ]
          }
        },
        // 不显示均值线
      }
    ]
  };
  return option;
}

function renderTable(data) {
  var labels = Object.keys(data);
  var tbody = document.querySelector('#data-table tbody');
  tbody.innerHTML = '';
  labels.forEach(function(k) {
    var row = document.createElement('tr');
    var td1 = document.createElement('td');
    td1.textContent = k;
    var td2 = document.createElement('td');
    td2.textContent = Number(data[k].dau).toLocaleString();
    var td3 = document.createElement('td');
    td3.textContent = data[k].community;
    row.appendChild(td1); row.appendChild(td2); row.appendChild(td3);
    tbody.appendChild(row);
  });
}

var currentData = daysData;
var chartDom = document.getElementById('chart');
var myChart = echarts.init(chartDom);
myChart.setOption(buildChart(daysData, 'daily'));
renderTable(daysData);

// Tab 切换
document.querySelectorAll('.tab-btn').forEach(function(btn) {
  btn.addEventListener('click', function() {
    document.querySelectorAll('.tab-btn').forEach(function(b) { b.classList.remove('active'); });
    btn.classList.add('active');
    var tab = btn.getAttribute('data-tab');
    var data, type;
    if (tab === 'daily') { data = daysData; type = 'daily'; }
    else if (tab === 'weekly') { data = weeksData; type = 'weekly'; }
    else { data = monthsData; type = 'monthly'; }

    // 更新表头
    var thead = document.querySelector('#data-table thead tr');
    if (tab === 'daily') { thead.innerHTML = '<th>日期</th><th>主动DAU</th><th>社区首页UV</th>'; }
    else if (tab === 'weekly') { thead.innerHTML = '<th>周期</th><th>主动DAU(周均)</th><th>社区首页UV(周均)</th>'; }
    else { thead.innerHTML = '<th>月份</th><th>主动DAU(月均)</th><th>社区首页UV(月均)</th>'; }

    myChart.setOption(buildChart(data, type));
    renderTable(data);
  });
});

window.addEventListener('resize', function() { myChart.resize(); });
</script>
</body>
</html>'''

html_path = os.path.join(save_dir, '主动DAU与社区首页UV_折线图.html')
with open(html_path, 'w', encoding='utf-8') as f:
    f.write(html)
print(f'HTML已生成: {html_path}')