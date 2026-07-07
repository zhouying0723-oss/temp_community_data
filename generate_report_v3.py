# -*- coding: utf-8 -*-
import os, json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ========== 原始数据（2025-01-01 ~ 2026-07-06） ==========
raw_data = [
    # 2025年1月
    ("2025-01-01", 544800, 1898), ("2025-01-02", 792680, 3526), ("2025-01-03", 789820, 3510),
    ("2025-01-04", 544473, 1891), ("2025-01-05", 413526, 1670), ("2025-01-06", 772581, 3313),
    ("2025-01-07", 799389, 3301), ("2025-01-08", 781760, 3449), ("2025-01-09", 770215, 3116),
    ("2025-01-10", 775461, 3293), ("2025-01-11", 544140, 1814), ("2025-01-12", 413743, 1568),
    ("2025-01-13", 756405, 3092), ("2025-01-14", 799650, 3304), ("2025-01-15", 778542, 3381),
    ("2025-01-16", 773145, 3159), ("2025-01-17", 772552, 2937), ("2025-01-18", 550685, 1719),
    ("2025-01-19", 457406, 1691), ("2025-01-20", 774518, 3150), ("2025-01-21", 800902, 3405),
    ("2025-01-22", 780916, 3232), ("2025-01-23", 771261, 3069), ("2025-01-24", 753028, 3027),
    ("2025-01-25", 542327, 1643), ("2025-01-26", 469767, 1504), ("2025-01-27", 696674, 2725),
    ("2025-01-28", 560615, 1866), ("2025-01-29", 543726, 1846), ("2025-01-30", 563296, 2110),
    ("2025-01-31", 546114, 1913),
    # 2025年2月
    ("2025-02-01", 534564, 1823), ("2025-02-02", 457178, 1600), ("2025-02-03", 607155, 2526),
    ("2025-02-04", 654005, 2644), ("2025-02-05", 812428, 3365), ("2025-02-06", 800234, 3217),
    ("2025-02-07", 802439, 3381), ("2025-02-08", 664698, 2276), ("2025-02-09", 471292, 1780),
    ("2025-02-10", 799566, 3199), ("2025-02-11", 820473, 3392), ("2025-02-12", 809134, 3283),
    ("2025-02-13", 817723, 3431), ("2025-02-14", 815810, 3355), ("2025-02-15", 604603, 2115),
    ("2025-02-16", 500232, 1985), ("2025-02-17", 820429, 3435), ("2025-02-18", 840217, 3522),
    ("2025-02-19", 828186, 3415), ("2025-02-20", 833199, 3562), ("2025-02-21", 833023, 3359),
    ("2025-02-22", 626342, 2276), ("2025-02-23", 499131, 2055), ("2025-02-24", 826740, 3486),
    ("2025-02-25", 845407, 3669), ("2025-02-26", 855935, 3697), ("2025-02-27", 850153, 3801),
    ("2025-02-28", 852779, 3947),
    # 2025年3月
    ("2025-03-01", 629843, 2227), ("2025-03-02", 500996, 2022), ("2025-03-03", 831663, 3603),
    ("2025-03-04", 875315, 3927), ("2025-03-05", 859738, 3775), ("2025-03-06", 870389, 3926),
    ("2025-03-07", 839589, 3750), ("2025-03-08", 623282, 2208), ("2025-03-09", 490196, 1892),
    ("2025-03-10", 821197, 3610), ("2025-03-11", 886003, 4121), ("2025-03-12", 871327, 3827),
    ("2025-03-13", 864193, 3836), ("2025-03-14", 864201, 3713), ("2025-03-15", 666877, 2178),
    ("2025-03-16", 517953, 1991), ("2025-03-17", 833089, 3489), ("2025-03-18", 864203, 3744),
    ("2025-03-19", 849583, 3679), ("2025-03-20", 852617, 4039), ("2025-03-21", 841864, 3559),
    ("2025-03-22", 607009, 2007), ("2025-03-23", 472357, 1780), ("2025-03-24", 819150, 3351),
    ("2025-03-25", 834182, 3432), ("2025-03-26", 828955, 3422), ("2025-03-27", 821771, 3615),
    ("2025-03-28", 825642, 3485), ("2025-03-29", 606867, 2031), ("2025-03-30", 470287, 1889),
    ("2025-03-31", 818190, 3626),
    # 2025年4月
    ("2025-04-01", 839150, 3589), ("2025-04-02", 837608, 3591), ("2025-04-03", 900847, 5145),
    ("2025-04-04", 769576, 4138), ("2025-04-05", 725568, 3357), ("2025-04-06", 585631, 2645),
    ("2025-04-07", 1089378, 7850), ("2025-04-08", 1066170, 6426), ("2025-04-09", 1095995, 7666),
    ("2025-04-10", 1110255, 7557), ("2025-04-11", 1055207, 6143), ("2025-04-12", 812221, 3580),
    ("2025-04-13", 619708, 2798), ("2025-04-14", 963291, 4495), ("2025-04-15", 959383, 4264),
    ("2025-04-16", 990177, 4821), ("2025-04-17", 967287, 4519), ("2025-04-18", 929618, 3684),
    ("2025-04-19", 609826, 2095), ("2025-04-20", 539904, 2132), ("2025-04-21", 930061, 4183),
    ("2025-04-22", 964980, 4634), ("2025-04-23", 988748, 5024), ("2025-04-24", 962250, 4651),
    ("2025-04-25", 928133, 4209), ("2025-04-26", 710264, 2565), ("2025-04-27", 612076, 2385),
    ("2025-04-28", 879764, 3777), ("2025-04-29", 878841, 3621), ("2025-04-30", 873457, 3490),
    # 2025年5月
    ("2025-05-01", 707474, 2664), ("2025-05-02", 708904, 2850), ("2025-05-03", 647456, 2230),
    ("2025-05-04", 503527, 1848), ("2025-05-05", 716065, 2984), ("2025-05-06", 919857, 3955),
    ("2025-05-07", 955578, 4458), ("2025-05-08", 944984, 4415), ("2025-05-09", 899014, 3981),
    ("2025-05-10", 679663, 2331), ("2025-05-11", 524601, 2112), ("2025-05-12", 977029, 4585),
    ("2025-05-13", 955694, 4222), ("2025-05-14", 932138, 4075), ("2025-05-15", 925377, 4347),
    ("2025-05-16", 888694, 3763), ("2025-05-17", 633570, 2093), ("2025-05-18", 514205, 1947),
    ("2025-05-19", 862671, 3656), ("2025-05-20", 868400, 3627), ("2025-05-21", 873972, 3587),
    ("2025-05-22", 896911, 3961), ("2025-05-23", 882903, 3828), ("2025-05-24", 622885, 2176),
    ("2025-05-25", 461296, 1900), ("2025-05-26", 834313, 3489), ("2025-05-27", 853914, 3539),
    ("2025-05-28", 848243, 3645), ("2025-05-29", 880549, 3750), ("2025-05-30", 847388, 3425),
    ("2025-05-31", 595283, 2125),
    # 2025年6月
    ("2025-06-01", 459227, 1896), ("2025-06-02", 702445, 3150), ("2025-06-03", 853938, 3681),
    ("2025-06-04", 849191, 3541), ("2025-06-05", 874423, 3933), ("2025-06-06", 851115, 3772),
    ("2025-06-07", 608733, 1994), ("2025-06-08", 495210, 2039), ("2025-06-09", 847706, 3578),
    ("2025-06-10", 892186, 3988), ("2025-06-11", 898663, 3855), ("2025-06-12", 909030, 3958),
    ("2025-06-13", 932060, 4370), ("2025-06-14", 703656, 2638), ("2025-06-15", 543914, 2429),
    ("2025-06-16", 886745, 3864), ("2025-06-17", 902488, 3892), ("2025-06-18", 905906, 4090),
    ("2025-06-19", 897639, 3991), ("2025-06-20", 864232, 3548), ("2025-06-21", 629359, 2138),
    ("2025-06-22", 617875, 2733), ("2025-06-23", 913440, 3957), ("2025-06-24", 945518, 4279),
    ("2025-06-25", 912103, 3652), ("2025-06-26", 890290, 3545), ("2025-06-27", 870761, 3400),
    ("2025-06-28", 633500, 2175), ("2025-06-29", 496299, 1884), ("2025-06-30", 837902, 3329),
    # 2025年7月
    ("2025-07-01", 863886, 3467), ("2025-07-02", 864392, 3317), ("2025-07-03", 872942, 3518),
    ("2025-07-04", 831134, 3028), ("2025-07-05", 575441, 1926), ("2025-07-06", 487199, 1943),
    ("2025-07-07", 823863, 3251), ("2025-07-08", 851472, 3238), ("2025-07-09", 860363, 3432),
    ("2025-07-10", 853590, 3296), ("2025-07-11", 845392, 3302), ("2025-07-12", 612286, 2047),
    ("2025-07-13", 478464, 1992), ("2025-07-14", 836722, 3406), ("2025-07-15", 860489, 3554),
    ("2025-07-16", 844619, 3421), ("2025-07-17", 836279, 3301), ("2025-07-18", 832237, 3283),
    ("2025-07-19", 591919, 1879), ("2025-07-20", 476490, 1975), ("2025-07-21", 831114, 3350),
    ("2025-07-22", 868184, 3641), ("2025-07-23", 871102, 3640), ("2025-07-24", 890843, 3819),
    ("2025-07-25", 845526, 3394), ("2025-07-26", 618687, 2193), ("2025-07-27", 503292, 2129),
    ("2025-07-28", 855486, 3450), ("2025-07-29", 861464, 3421), ("2025-07-30", 884152, 3782),
    ("2025-07-31", 890909, 3903),
    # 2025年8月
    ("2025-08-01", 870937, 3663), ("2025-08-02", 647389, 2324), ("2025-08-03", 496184, 2065),
    ("2025-08-04", 835752, 3383), ("2025-08-05", 847853, 3419), ("2025-08-06", 850350, 3283),
    ("2025-08-07", 849439, 3424), ("2025-08-08", 823479, 3249), ("2025-08-09", 596885, 2112),
    ("2025-08-10", 478916, 2045), ("2025-08-11", 837782, 3367), ("2025-08-12", 863340, 3637),
    ("2025-08-13", 865620, 3587), ("2025-08-14", 877329, 3703), ("2025-08-15", 862107, 3513),
    ("2025-08-16", 624259, 2417), ("2025-08-17", 476547, 2096), ("2025-08-18", 851521, 3445),
    ("2025-08-19", 881467, 3683), ("2025-08-20", 887553, 3855), ("2025-08-21", 878944, 3810),
    ("2025-08-22", 898361, 4150), ("2025-08-23", 668290, 2806), ("2025-08-24", 523310, 2470),
    ("2025-08-25", 898746, 4197), ("2025-08-26", 917737, 4197), ("2025-08-27", 914524, 4297),
    ("2025-08-28", 923612, 4250), ("2025-08-29", 897993, 3966), ("2025-08-30", 638405, 2592),
    ("2025-08-31", 536868, 2649),
    # 2025年9月
    ("2025-09-01", 888452, 3921), ("2025-09-02", 927773, 4355), ("2025-09-03", 910536, 4290),
    ("2025-09-04", 920891, 4371), ("2025-09-05", 900776, 3996), ("2025-09-06", 674594, 2688),
    ("2025-09-07", 529666, 2495), ("2025-09-08", 901207, 3969), ("2025-09-09", 927193, 4202),
    ("2025-09-10", 934459, 4090), ("2025-09-11", 944093, 4486), ("2025-09-12", 919046, 4055),
    ("2025-09-13", 648964, 2610), ("2025-09-14", 520527, 2570), ("2025-09-15", 918582, 4030),
    ("2025-09-16", 938998, 4272), ("2025-09-17", 967975, 4749), ("2025-09-18", 1024179, 5637),
    ("2025-09-19", 918496, 3970), ("2025-09-20", 632552, 2408), ("2025-09-21", 509597, 2377),
    ("2025-09-22", 914607, 4019), ("2025-09-23", 938182, 4223), ("2025-09-24", 928479, 4256),
    ("2025-09-25", 926854, 4182), ("2025-09-26", 909661, 3884), ("2025-09-27", 627317, 2459),
    ("2025-09-28", 569769, 2382), ("2025-09-29", 864204, 3588), ("2025-09-30", 878949, 3493),
    # 2025年10月
    ("2025-10-01", 653225, 2518), ("2025-10-02", 667781, 2725), ("2025-10-03", 632253, 2542),
    ("2025-10-04", 570260, 2012), ("2025-10-05", 484894, 1964), ("2025-10-06", 622794, 2566),
    ("2025-10-07", 663633, 2726), ("2025-10-08", 720231, 3143), ("2025-10-09", 941960, 4168),
    ("2025-10-10", 952142, 4661), ("2025-10-11", 866403, 4285), ("2025-10-12", 661836, 3214),
    ("2025-10-13", 1081909, 5285), ("2025-10-14", 1100392, 5422), ("2025-10-15", 1102014, 4980),
    ("2025-10-16", 1089127, 4711), ("2025-10-17", 1100875, 5363), ("2025-10-18", 848155, 3244),
    ("2025-10-19", 688822, 2797), ("2025-10-20", 1060774, 4374), ("2025-10-21", 1094995, 4770),
    ("2025-10-22", 1112008, 4840), ("2025-10-23", 1105935, 4836), ("2025-10-24", 1081041, 4426),
    ("2025-10-25", 808533, 2668), ("2025-10-26", 688233, 2766), ("2025-10-27", 1103886, 4549),
    ("2025-10-28", 1114525, 4813), ("2025-10-29", 1112246, 4860), ("2025-10-30", 1155860, 5797),
    ("2025-10-31", 1089363, 4396),
    # 2025年11月
    ("2025-11-01", 803129, 2702), ("2025-11-02", 644713, 2392), ("2025-11-03", 1064840, 4214),
    ("2025-11-04", 1082312, 4338), ("2025-11-05", 1090187, 4612), ("2025-11-06", 1079802, 4371),
    ("2025-11-07", 1071894, 4355), ("2025-11-08", 782244, 2668), ("2025-11-09", 650187, 2447),
    ("2025-11-10", 1055984, 4084), ("2025-11-11", 1071106, 4172), ("2025-11-12", 1066873, 4166),
    ("2025-11-13", 1081003, 4566), ("2025-11-14", 1027351, 4829), ("2025-11-15", 774430, 2822),
    ("2025-11-16", 604441, 2487), ("2025-11-17", 1007411, 4338), ("2025-11-18", 1036338, 4509),
    ("2025-11-19", 1032019, 4503), ("2025-11-20", 1032227, 4805), ("2025-11-21", 1059121, 5147),
    ("2025-11-22", 764400, 2800), ("2025-11-23", 585242, 2476), ("2025-11-24", 1011935, 4189),
    ("2025-11-25", 1051744, 4506), ("2025-11-26", 1029347, 4185), ("2025-11-27", 1004148, 3843),
    ("2025-11-28", 979585, 3724), ("2025-11-29", 736993, 2576), ("2025-11-30", 586155, 2298),
    # 2025年12月
    ("2025-12-01", 990914, 4054), ("2025-12-02", 1012612, 4254), ("2025-12-03", 1010840, 4373),
    ("2025-12-04", 1006405, 4042), ("2025-12-05", 981577, 3749), ("2025-12-06", 723677, 2550),
    ("2025-12-07", 560746, 2252), ("2025-12-08", 979486, 3783), ("2025-12-09", 1001761, 4081),
    ("2025-12-10", 1012719, 4218), ("2025-12-11", 1034741, 4788), ("2025-12-12", 994479, 4177),
    ("2025-12-13", 762590, 2844), ("2025-12-14", 571768, 2292), ("2025-12-15", 971913, 4154),
    ("2025-12-16", 1020486, 4667), ("2025-12-17", 1013829, 4344), ("2025-12-18", 1018149, 4680),
    ("2025-12-19", 1000679, 4295), ("2025-12-20", 727651, 2554), ("2025-12-21", 550088, 2352),
    ("2025-12-22", 974857, 4273), ("2025-12-23", 1016848, 4640), ("2025-12-24", 1012670, 4760),
    ("2025-12-25", 987369, 4259), ("2025-12-26", 975070, 4513), ("2025-12-27", 736971, 3320),
    ("2025-12-28", 582533, 2933), ("2025-12-29", 1022159, 5602), ("2025-12-30", 1032078, 5284),
    ("2025-12-31", 984399, 4537),
    # 2026年1月
    ("2026-01-01", 697434, 2671), ("2026-01-02", 738751, 3367), ("2026-01-03", 763515, 3456),
    ("2026-01-04", 777600, 3539), ("2026-01-05", 1050682, 4920), ("2026-01-06", 1064749, 5003),
    ("2026-01-07", 1064528, 5242), ("2026-01-08", 1055436, 5216), ("2026-01-09", 1049381, 5024),
    ("2026-01-10", 783241, 3538), ("2026-01-11", 618151, 3065), ("2026-01-12", 1077513, 5282),
    ("2026-01-13", 1107598, 5708), ("2026-01-14", 1116877, 5823), ("2026-01-15", 1117369, 5957),
    ("2026-01-16", 1061788, 5184), ("2026-01-17", 782239, 3332), ("2026-01-18", 639477, 3088),
    ("2026-01-19", 1039566, 4897), ("2026-01-20", 1080996, 5209), ("2026-01-21", 1097765, 5843),
    ("2026-01-22", 1090907, 5831), ("2026-01-23", 1081513, 5975), ("2026-01-24", 819534, 4080),
    ("2026-01-25", 648203, 3577), ("2026-01-26", 1089991, 6180), ("2026-01-27", 1110726, 6919),
    ("2026-01-28", 1160277, 7889), ("2026-01-29", 1228498, 10498), ("2026-01-30", 1262245, 12674),
    ("2026-01-31", 1061808, 9406),
    # 2026年2月
    ("2026-02-01", 834449, 6280), ("2026-02-02", 1317266, 12382), ("2026-02-03", 1276626, 9864),
    ("2026-02-04", 1268744, 9345), ("2026-02-05", 1262299, 9713), ("2026-02-06", 1228655, 8675),
    ("2026-02-07", 959856, 4773), ("2026-02-08", 704573, 3529), ("2026-02-09", 1148995, 6321),
    ("2026-02-10", 1148909, 6006), ("2026-02-11", 1136855, 5870), ("2026-02-12", 1108698, 5171),
    ("2026-02-13", 1087050, 5478), ("2026-02-14", 836545, 3099), ("2026-02-15", 628715, 2317),
    ("2026-02-16", 739366, 2711), ("2026-02-17", 785085, 3106), ("2026-02-18", 785266, 3110),
    ("2026-02-19", 781480, 3235), ("2026-02-20", 804540, 3451), ("2026-02-21", 790059, 3086),
    ("2026-02-22", 640467, 2391), ("2026-02-23", 903843, 4369), ("2026-02-24", 1162268, 5748),
    ("2026-02-25", 1167330, 5453), ("2026-02-26", 1174942, 5961), ("2026-02-27", 1183555, 5895),
    ("2026-02-28", 1161031, 6225),
    # 2026年3月
    ("2026-03-01", 1013315, 5611), ("2026-03-02", 1347844, 8387), ("2026-03-03", 1351006, 8262),
    ("2026-03-04", 1341788, 8010), ("2026-03-05", 1299978, 7139), ("2026-03-06", 1274324, 6838),
    ("2026-03-07", 1004620, 4472), ("2026-03-08", 783047, 3627), ("2026-03-09", 1287956, 7218),
    ("2026-03-10", 1294487, 6945), ("2026-03-11", 1288497, 6632), ("2026-03-12", 1285044, 6279),
    ("2026-03-13", 1268924, 6259), ("2026-03-14", 982524, 4376), ("2026-03-15", 801533, 3737),
    ("2026-03-16", 1267504, 6293), ("2026-03-17", 1270361, 5864), ("2026-03-18", 1286717, 6561),
    ("2026-03-19", 1347459, 8092), ("2026-03-20", 1305185, 6796), ("2026-03-21", 1005247, 4688),
    ("2026-03-22", 842322, 4072), ("2026-03-23", 1431359, 9629), ("2026-03-24", 1402543, 7913),
    ("2026-03-25", 1362207, 6915), ("2026-03-26", 1337357, 6432), ("2026-03-27", 1323689, 6018),
    ("2026-03-28", 1024829, 4035), ("2026-03-29", 810717, 3380), ("2026-03-30", 1294899, 5926),
    ("2026-03-31", 1305405, 5518),
    # 2026年4月
    ("2026-04-01", 1324127, 5994), ("2026-04-02", 1331661, 6306), ("2026-04-03", 1266079, 4905),
    ("2026-04-04", 857954, 2961), ("2026-04-05", 747474, 2839), ("2026-04-06", 1015920, 4259),
    ("2026-04-07", 1306284, 5524), ("2026-04-08", 1409776, 6730), ("2026-04-09", 1352125, 5922),
    ("2026-04-10", 1297181, 5038), ("2026-04-11", 987708, 3490), ("2026-04-12", 905167, 3873),
    ("2026-04-13", 1311431, 5479), ("2026-04-14", 1295368, 5018), ("2026-04-15", 1292279, 4893),
    ("2026-04-16", 1280175, 4656), ("2026-04-17", 1284058, 4877), ("2026-04-18", 1039217, 3582),
    ("2026-04-19", 817110, 3125), ("2026-04-20", 1272666, 4847), ("2026-04-21", 1284351, 4901),
    ("2026-04-22", 1270675, 4877), ("2026-04-23", 1267728, 5031), ("2026-04-24", 1263222, 5043),
    ("2026-04-25", 916309, 2871), ("2026-04-26", 768172, 2690), ("2026-04-27", 1208252, 4610),
    ("2026-04-28", 1240181, 4630), ("2026-04-29", 1234415, 4429), ("2026-04-30", 1202546, 4155),
    # 2026年5月
    ("2026-05-01", 906841, 2796), ("2026-05-02", 815379, 2263), ("2026-05-03", 670309, 1934),
    ("2026-05-04", 900993, 2981), ("2026-05-05", 998869, 3227), ("2026-05-06", 1256736, 4536),
    ("2026-05-07", 1250968, 4433), ("2026-05-08", 1248021, 4441), ("2026-05-09", 998482, 2637),
    ("2026-05-10", 734794, 2267), ("2026-05-11", 1230500, 4124), ("2026-05-12", 1244904, 4385),
    ("2026-05-13", 1284565, 4407), ("2026-05-14", 1295919, 4384), ("2026-05-15", 1286894, 4487),
    ("2026-05-16", 923563, 2699), ("2026-05-17", 747602, 2347), ("2026-05-18", 1234690, 4376),
    ("2026-05-19", 1236846, 4323), ("2026-05-20", 1254695, 4392), ("2026-05-21", 1267352, 4646),
    ("2026-05-22", 1235488, 4091), ("2026-05-23", 947186, 2554), ("2026-05-24", 768694, 2619),
    ("2026-05-25", 1201080, 3989), ("2026-05-26", 1227753, 4107), ("2026-05-27", 1249255, 4459),
    ("2026-05-28", 1238903, 4430), ("2026-05-29", 1225926, 4072), ("2026-05-30", 872072, 2226),
    ("2026-05-31", 663601, 1992),
    # 2026年6月
    ("2026-06-01", 1182600, 3801), ("2026-06-02", 1213861, 3961), ("2026-06-03", 1213703, 4152),
    ("2026-06-04", 1212854, 4168), ("2026-06-05", 1211218, 4334), ("2026-06-06", 965831, 3009),
    ("2026-06-07", 722395, 2354), ("2026-06-08", 1261710, 4822), ("2026-06-09", 1257023, 4419),
    ("2026-06-10", 1310748, 5763), ("2026-06-11", 1320599, 5317), ("2026-06-12", 1321505, 5104),
    ("2026-06-13", 986432, 2705), ("2026-06-14", 743602, 2309), ("2026-06-15", 1266008, 4509),
    ("2026-06-16", 1267815, 4527), ("2026-06-17", 1255257, 4303), ("2026-06-18", 1259098, 4537),
    ("2026-06-19", 1025799, 3262), ("2026-06-20", 829099, 2309), ("2026-06-21", 760663, 2373),
    ("2026-06-22", 1235126, 4103), ("2026-06-23", 1276774, 4475), ("2026-06-24", 1290074, 4518),
    ("2026-06-25", 1289016, 4801), ("2026-06-26", 1266258, 4397), ("2026-06-27", 961133, 2555),
    ("2026-06-28", 744261, 2356), ("2026-06-29", 1216927, 4112), ("2026-06-30", 1240069, 3928),
    # 2026年7月
    ("2026-07-01", 1257607, 4248), ("2026-07-02", 1278013, 4585), ("2026-07-03", 1237187, 3918),
    ("2026-07-04", 834422, 1985), ("2026-07-05", 651250, 1907), ("2026-07-06", 1196419, 3777),
]

dates_full = [d[0] for d in raw_data]
dau_full = [d[1] for d in raw_data]
community_full = [d[2] for d in raw_data]

# ========== 周均计算（每7天一周，不足7天丢弃） ==========
def calc_weekly(dates, dau_list, com_list):
    weeks = []
    week_labels = []
    for i in range(0, len(dates), 7):
        if i + 7 > len(dates):
            break
        week = dates[i:i+7]
        week_dau_avg = round(sum(dau_list[i:i+7]) / 7)
        week_com_avg = round(sum(com_list[i:i+7]) / 7)
        # 格式: 251201~251207
        sd = week[0][2:4] + week[0][5:7] + week[0][8:10]
        ed = week[6][5:7] + week[6][8:10]
        weeks.append({'dau': week_dau_avg, 'community': week_com_avg})
        week_labels.append(f"{sd}~{ed}")
    return week_labels, weeks

week_labels, weekly_data = calc_weekly(dates_full, dau_full, community_full)
weeks_dict = {}
for i, wl in enumerate(week_labels):
    weeks_dict[wl] = weekly_data[i]

# ========== 月均 ==========
monthly_dict = {}
for d, dd, dc in zip(dates_full, dau_full, community_full):
    m = d[:7]
    if m not in monthly_dict:
        monthly_dict[m] = {'dau_list': [], 'com_list': []}
    monthly_dict[m]['dau_list'].append(dd)
    monthly_dict[m]['com_list'].append(dc)

months_dict = {}
for m in sorted(monthly_dict.keys()):
    d_list = monthly_dict[m]['dau_list']
    c_list = monthly_dict[m]['com_list']
    months_dict[m] = {'dau': round(sum(d_list)/len(d_list)), 'community': round(sum(c_list)/len(c_list))}

save_dir = os.path.dirname(os.path.abspath(__file__))

# ========== 保存 Excel ==========
wb = openpyxl.Workbook()
header_font = Font(bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def write_header(ws, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border

# Sheet1: 日数据
ws1 = wb.active; ws1.title = '日数据'
write_header(ws1, ['日期', '主动DAU', '社区首页UV'])
for i, (d, dau, cu) in enumerate(raw_data):
    ws1.cell(row=i+2, column=1, value=d).border = thin_border
    ws1.cell(row=i+2, column=1).alignment = center_align
    ws1.cell(row=i+2, column=2, value=dau).border = thin_border
    ws1.cell(row=i+2, column=3, value=cu).border = thin_border
ws1.column_dimensions['A'].width = 14; ws1.column_dimensions['B'].width = 14; ws1.column_dimensions['C'].width = 14

# Sheet2: 周均
ws2 = wb.create_sheet('周均')
write_header(ws2, ['周期', '主动DAU(周均)', '社区首页UV(周均)'])
for i, wl in enumerate(week_labels):
    ws2.cell(row=i+2, column=1, value=wl).border = thin_border
    ws2.cell(row=i+2, column=1).alignment = center_align
    ws2.cell(row=i+2, column=2, value=weekly_data[i]['dau']).border = thin_border
    ws2.cell(row=i+2, column=3, value=weekly_data[i]['community']).border = thin_border
ws2.column_dimensions['A'].width = 14; ws2.column_dimensions['B'].width = 16; ws2.column_dimensions['C'].width = 18

# Sheet3: 月均
ws3 = wb.create_sheet('月均')
write_header(ws3, ['月份', '主动DAU(月均)', '社区首页UV(月均)'])
for i, m in enumerate(sorted(months_dict.keys())):
    ws3.cell(row=i+2, column=1, value=m).border = thin_border
    ws3.cell(row=i+2, column=1).alignment = center_align
    ws3.cell(row=i+2, column=2, value=months_dict[m]['dau']).border = thin_border
    ws3.cell(row=i+2, column=3, value=months_dict[m]['community']).border = thin_border
ws3.column_dimensions['A'].width = 10; ws3.column_dimensions['B'].width = 16; ws3.column_dimensions['C'].width = 18

xlsx_path = os.path.join(save_dir, '主动DAU与社区首页UV数据.xlsx')
wb.save(xlsx_path)
print(f'Excel已保存: {xlsx_path}')

# ========== 计算功能渗透率 ==========
def calc_penetration(dau_val, com_val):
    """功能渗透率 = 社区首页UV / 主动DAU，返回百分比"""
    if dau_val == 0:
        return 0
    return round(com_val / dau_val * 100, 2)

# 日渗透率
days_penetration = {}
for d, dau, cu in zip(dates_full, dau_full, community_full):
    days_penetration[d] = round(cu / dau * 100, 2) if dau > 0 else 0

# 周渗透率(周均)
weeks_penetration = {}
for wl, wd in zip(week_labels, weekly_data):
    weeks_penetration[wl] = round(wd['community'] / wd['dau'] * 100, 2) if wd['dau'] > 0 else 0

# 月渗透率(月均)
months_penetration = {}
for m in sorted(months_dict.keys()):
    d = months_dict[m]
    months_penetration[m] = round(d['community'] / d['dau'] * 100, 2) if d['dau'] > 0 else 0

# ========== 日数据字典 ==========
days_dict = {}
for d, dau, cu in zip(dates_full, dau_full, community_full):
    days_dict[d] = {'dau': dau, 'community': cu}

# ========== 生成 ECharts HTML ==========
# 所有标注（喜娜AI / 社区/股吧/评论相关改动）
all_annotations = [
    # (日期, 描述, 标签)
    # 喜娜AI
    ("2024-05-29", "喜娜AI:正文页AI帮抢沙发&评论优化", "xina"),
    ("2024-06-11", "喜娜AI:正文页AI摘要与语音播报合并到一行", "xina"),
    ("2024-07-01", "喜娜AI:快讯详情页补帮写数据埋点", "xina"),
    ("2025-04-15", "喜娜AI:前台页面相关接口", "xina"),
    ("2025-04-25", "喜娜AI:股吧回复", "xina"),
    ("2025-12-15", "喜娜AI入口更换为芝麻AI", "xina"),
    ("2026-01-30", "鸿蒙:芝麻喜娜AI摘要改为芝麻AI摘要", "xina"),
    ("2026-03-26", "鸿蒙:正文页支持喜娜提取摘要", "xina"),
    # 社区/股吧/评论
    ("2025-03-19", "个股评论右上互动消息入口改为消息中心入口", "community"),
    ("2025-03-26", "分享到社区:发帖支持仅传图&展示schema", "community"),
    ("2025-03-27", "发帖优化(股吧):帖子不展示在个人主页", "community"),
    ("2025-04-15", "股票评论数据打点stock_comment统一", "community"),
    ("2025-04-16", "社区首页返回按钮", "community"),
    ("2025-04-23", "消息中心:清除未读数交互优化", "community"),
    ("2025-04-29", "社区股吧帖子分类页卡(股吧tab)", "community"),
    ("2025-05-07", "期货社区讨论吧", "community"),
    ("2025-05-19", "社区-投票分享改为长图", "community"),
    ("2025-05-21", "社区-主页同步账号视频内容", "community"),
    ("2025-06-05", "社区导航增加热议个股/新闻分享到社区挂链", "community"),
    ("2025-06-06", "讨论吧帖子加精样式", "community"),
    ("2025-06-17", "社区发帖编辑界面图片位置大小优化", "community"),
    ("2025-07-07", "新闻评论进搜索", "community"),
    ("2025-07-11", "新闻评论@进消息信箱&评论详情页&分享海报", "community"),
    ("2025-07-14", "我的关注-股吧及新闻帖子增加分享长图海报", "community"),
    ("2025-07-25", "新闻评论长图分享海报位置补充", "community"),
    ("2025-08-21", "个股评论列表页最新发帖/最新回复移到一级tab", "community"),
    ("2025-08-29", "新闻评论中个股相关评论进股吧", "community"),
    ("2025-09-01", "新闻/个股评论详情页回复顺序优化", "community"),
    ("2025-09-15", "讨论吧聚合个股评论中的新闻评论", "community"),
    ("2025-09-19", "个股评论列表页发主贴改为全屏发帖", "community"),
    ("2025-10-14", "社区公约", "community"),
    ("2025-10-15", "点击新闻和724评论头像进个人主页", "community"),
    ("2025-11-06", "社区付费投票(投票/支付/个人主页/徽章)", "community"),
    ("2025-11-25", "个人主页帖子列表&评论发布后显示徽章/炫彩昵称/头像框", "community"),
    ("2026-03-09", "互动消息增加新闻评论", "community"),
    ("2026-03-26", "社区及行情评论-语音消息", "community"),
    ("2026-04-20", "社区-吧内支持屏蔽发帖功能", "community"),
    ("2026-04-21", "纽约原油评论页tab支持scheme", "community"),
    ("2026-04-29", "社区-大V私享会相关需求", "community"),
    ("2026-05-06", "社区-评论送票", "community"),
    ("2026-05-19", "社区-投票分享改为长图", "community"),
    ("2026-05-29", "社区-发帖图片查看组件分享按钮去除", "community"),
]

# 生成JS可用的标注数据（按日期合并，同一天可能有多个标签）
# 每个标注对象: {date, label, tag, show: true}
annotations_js = [{"date": a[0], "label": a[1], "tag": a[2]} for a in all_annotations]
html = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>主动DAU & 社区首页UV</title>
<script src="https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js"></script>
<style>
* { margin: 0; padding: 0; box-sizing: border-box; }
body { background: #f5f7fa; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; padding: 20px; }
.container { max-width: 1100px; margin: 0 auto; background: #fff; border-radius: 12px; padding: 30px; }
.chart-section { margin-bottom: 10px; }
h2 { text-align: center; color: #333; font-weight: 500; margin-bottom: 20px; }
h3 { text-align: center; color: #555; font-weight: 500; margin: 20px 0 15px 0; font-size: 16px; }
h4 { color: #444; font-weight: 500; margin: 20px 0 10px 0; font-size: 14px; }
.tabs { display: flex; justify-content: center; gap: 10px; margin-bottom: 20px; }
.tab-btn { padding: 8px 24px; border: 1px solid #d0d5dd; border-radius: 6px; background: #fff; cursor: pointer; font-size: 14px; color: #555; transition: all .2s; }
.tab-btn:hover { border-color: #5470C6; color: #5470C6; }
.tab-btn.active { background: #5470C6; color: #fff; border-color: #5470C6; }
#chart { width: 100%; height: 420px; }
#chart2 { width: 100%; height: 350px; }
.table-wrap { margin-top: 20px; overflow-x: auto; max-height: 350px; overflow-y: auto; }
.separator { border: none; border-top: 1px solid #e8e8e8; margin: 25px 0; }
table { width: 100%; border-collapse: collapse; font-size: 12px; }
th { background: #f0f4ff; color: #333; font-weight: 600; padding: 8px 10px; border: 1px solid #e0e5ee; text-align: center; position: sticky; top: 0; z-index: 1; }
td { padding: 6px 10px; border: 1px solid #e0e5ee; text-align: center; color: #444; }
tr:nth-child(even) { background: #fafbfc; }
tr:hover { background: #f0f4ff; }
.info { text-align: center; color: #999; font-size: 12px; margin-top: 12px; }
.annotation-list { margin-top: 10px; max-height: 400px; overflow-y: auto; border: 1px solid #e8e8e8; border-radius: 8px; padding: 10px 14px; }
.annotation-item { display: flex; align-items: center; gap: 10px; padding: 5px 4px; border-bottom: 1px solid #f0f0f0; font-size: 13px; line-height: 1.5; }
.annotation-item:last-child { border-bottom: none; }
.annotation-item input[type="checkbox"] { cursor: pointer; flex-shrink: 0; width: 16px; height: 16px; }
.tag-xina { display: inline-block; background: #3498DB; color: #fff; font-size: 11px; padding: 0 8px; border-radius: 10px; flex-shrink: 0; }
.tag-community { display: inline-block; background: #27AE60; color: #fff; font-size: 11px; padding: 0 8px; border-radius: 10px; flex-shrink: 0; }
.ann-date { color: #888; font-size: 12px; flex-shrink: 0; width: 82px; }
.ann-label { color: #333; flex: 1; }
.legend-row { display: flex; align-items: center; gap: 20px; margin-bottom: 12px; font-size: 13px; }
.legend-dot { display: inline-block; width: 12px; height: 3px; border-radius: 2px; margin-right: 4px; }
</style>
</head>
<body>
<div class="container">
  <div class="tabs">
    <button class="tab-btn active" data-tab="daily">日数据</button>
    <button class="tab-btn" data-tab="weekly">周均</button>
    <button class="tab-btn" data-tab="monthly">月均</button>
  </div>

  <h2>DAU & 社区首页UV</h2>
  <div id="chart"></div>
  <div class="table-wrap">
    <table id="data-table">
      <thead><tr><th>日期</th><th>主动DAU</th><th>社区首页UV</th></tr></thead>
      <tbody></tbody>
    </table>
  </div>

  <hr class="separator">

  <h3>社区首页功能渗透率</h3>
  <div id="chart2"></div>
  <div class="table-wrap">
    <table id="penetration-table">
      <thead><tr><th>日期</th><th>功能渗透率</th></tr></thead>
      <tbody></tbody>
    </table>
  </div>

  <hr class="separator">

  <h3>需求标注清单</h3>
  <div class="legend-row">
    <span><span class="legend-dot" style="background:#3498DB;"></span>喜娜AI</span>
    <span><span class="legend-dot" style="background:#27AE60;"></span>社区/股吧/评论</span>
    <span style="color:#999;">勾选开关: 控制图上对应竖线显隐</span>
  </div>
  <div class="annotation-list" id="annotationList"></div>

  <div class="info">数据周期: 2025-01-01 ~ 2026-07-06</div>
</div>
<script>
var daysData = ''' + json.dumps(days_dict) + ''';
var weeksData = ''' + json.dumps(weeks_dict) + ''';
var monthsData = ''' + json.dumps(months_dict) + ''';

var daysPenetration = ''' + json.dumps(days_penetration) + ''';
var weeksPenetration = ''' + json.dumps(weeks_penetration) + ''';
var monthsPenetration = ''' + json.dumps(months_penetration) + ''';

var allAnnotations = ''' + json.dumps(annotations_js) + ''';

// 每个标注默认显示
var showMap = {};
allAnnotations.forEach(function(a, i) { showMap[i] = true; });

// 标注颜色: xina -> 蓝色, community -> 绿色
var tagColors = { xina: '#3498DB', community: '#27AE60' };
var tagLabels = { xina: '喜娜AI', community: '社区' };

// 获取当前应显示的markLine数据
function getMarkLines(type) {
  var markers = [];
  allAnnotations.forEach(function(a, i) {
    if (!showMap[i]) return;
    var color = tagColors[a.tag];
    var label = type === 'daily' ? a.label.split(':')[0] : '';
    markers.push({
      xAxis: a.date,
      label: { formatter: label, fontSize: 9, color: color, position: 'start' },
      lineStyle: { color: color, type: 'dashed', width: 1 }
    });
  });
  return markers;
}

// 获取tooltip中显示的信息
function getAnnotationInfo(key) {
  var parts = [];
  allAnnotations.forEach(function(a) {
    if (a.date === key) {
      var color = tagColors[a.tag];
      parts.push('<div style="color:' + color + ';margin:2px 0;">[' + tagLabels[a.tag] + '] ' + a.label + '</div>');
    }
  });
  if (parts.length > 0) {
    return '<div style="margin-top:6px;padding-top:6px;border-top:1px dashed #e0e0e0;">' + parts.join('') + '</div>';
  }
  return '';
}

function buildChart(data, type) {
  var labels = Object.keys(data);
  var dauVals = labels.map(function(k) { return data[k].dau; });
  var comVals = labels.map(function(k) { return data[k].community; });

  var option = {
    tooltip: {
      trigger: 'axis',
      axisPointer: { type: 'cross' },
      backgroundColor: 'rgba(255,255,255,0.95)',
      borderColor: '#e0e0e0', borderWidth: 1, padding: [10, 14],
      formatter: function(params) {
        var key = params[0].axisValue;
        var s = '<div style="font-weight:600;margin-bottom:6px;">' + key + '</div>';
        params.forEach(function(p) {
          var v = p.seriesIndex === 0 ? Number(p.value).toLocaleString() : p.value;
          s += '<div style="display:flex;align-items:center;gap:6px;margin:3px 0;">' +
            '<span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:' + p.color + ';"></span>' +
            p.seriesName + ': <strong>' + v + '</strong></div>';
        });
        return s + getAnnotationInfo(key);
      }
    },
    legend: { data: ['主动DAU', '社区首页UV'], top: 5, textStyle: { fontSize: 13 } },
    grid: { left: 60, right: 60, bottom: 40, top: 50 },
    xAxis: {
      type: 'category', data: labels,
      axisLabel: { fontSize: 10, rotate: type === 'daily' ? 45 : 0, interval: type === 'daily' ? 13 : 0 },
      axisLine: { lineStyle: { color: '#ccc' } }
    },
    yAxis: [
      {
        type: 'value', name: '主动DAU',
        nameTextStyle: { fontSize: 13, fontWeight: 'bold', padding: [0,0,0,40] },
        min: 0,
        splitLine: { lineStyle: { color: '#f0f0f0', type: 'dashed' } },
        axisLabel: { fontSize: 11, formatter: function(v) { return v >= 10000 ? (v/10000).toFixed(0) + '万' : v; } }
      },
      {
        type: 'value', name: '社区首页UV',
        nameTextStyle: { fontSize: 13, fontWeight: 'bold', padding: [0,40,0,0] },
        min: 0,
        splitLine: { show: false },
        axisLabel: { fontSize: 11 }
      }
    ],
    series: [
      {
        name: '主动DAU', type: 'line', yAxisIndex: 0,
        data: dauVals, smooth: true,
        symbol: 'none',
        lineStyle: { width: 2, color: '#5470C6' },
        itemStyle: { color: '#5470C6' },
        areaStyle: {
          color: { type: 'linear', x: 0, y: 0, x2: 0, y2: 1,
            colorStops: [
              { offset: 0, color: 'rgba(84,112,198,0.25)' },
              { offset: 1, color: 'rgba(84,112,198,0.02)' }
            ]
          }
        },
        markLine: {
          silent: true,
          symbol: ['none', 'none'],
          data: getMarkLines(type),
          label: { fontSize: 9 },
          lineStyle: { type: 'dashed', width: 1 }
        }
      },
      {
        name: '社区首页UV', type: 'line', yAxisIndex: 1,
        data: comVals, smooth: true,
        symbol: 'none',
        lineStyle: { width: 2, color: '#EE6666' },
        itemStyle: { color: '#EE6666' },
        areaStyle: {
          color: { type: 'linear', x: 0, y: 0, x2: 0, y2: 1,
            colorStops: [
              { offset: 0, color: 'rgba(238,102,102,0.25)' },
              { offset: 1, color: 'rgba(238,102,102,0.02)' }
            ]
          }
        },
        markLine: {
          silent: true,
          symbol: ['none', 'none'],
          data: getMarkLines(type),
          label: { fontSize: 9 },
          lineStyle: { type: 'dashed', width: 1 }
        }
      }
    ]
  };
  return option;
}

function buildPenetrationChart(data, type) {
  var labels = Object.keys(data);
  var vals = labels.map(function(k) { return data[k]; });

  var option = {
    tooltip: {
      trigger: 'axis',
      axisPointer: { type: 'cross' },
      backgroundColor: 'rgba(255,255,255,0.95)',
      borderColor: '#e0e0e0', borderWidth: 1, padding: [10, 14],
      formatter: function(params) {
        var p = params[0];
        return '<div style="font-weight:600;margin-bottom:6px;">' + p.axisValue + '</div>' +
          '功能渗透率: <strong>' + p.value + '%</strong>' +
          getAnnotationInfo(p.axisValue);
      }
    },
    grid: { left: 60, right: 30, bottom: 40, top: 30 },
    xAxis: {
      type: 'category', data: labels,
      axisLabel: { fontSize: 10, rotate: type === 'daily' ? 45 : 0, interval: type === 'daily' ? 13 : 0 },
      axisLine: { lineStyle: { color: '#ccc' } }
    },
    yAxis: {
      type: 'value', name: '渗透率(%)',
      nameTextStyle: { fontSize: 12, fontWeight: 'bold' },
      splitLine: { lineStyle: { color: '#f0f0f0', type: 'dashed' } },
      axisLabel: { fontSize: 11, formatter: '{value}%' }
    },
    series: [
      {
        name: '功能渗透率', type: 'line',
        data: vals, smooth: true,
        symbol: type === 'daily' ? 'none' : 'circle',
        symbolSize: 6,
        lineStyle: { width: 3, color: '#91CC75' },
        itemStyle: { color: '#91CC75' },
        areaStyle: {
          color: { type: 'linear', x: 0, y: 0, x2: 0, y2: 1,
            colorStops: [
              { offset: 0, color: 'rgba(145,204,117,0.35)' },
              { offset: 1, color: 'rgba(145,204,117,0.02)' }
            ]
          }
        },
        markLine: {
          silent: true,
          symbol: ['none', 'none'],
          data: getMarkLines(type),
          label: { fontSize: 9 },
          lineStyle: { type: 'dashed', width: 1 }
        }
      }
    ]
  };
  return option;
}

function renderTable(data, tab, tableId) {
  var labels = Object.keys(data);
  var tbody = document.querySelector('#' + tableId + ' tbody');
  tbody.innerHTML = '';
  labels.forEach(function(k) {
    var row = document.createElement('tr');
    var td1 = document.createElement('td'); td1.textContent = k;
    var val = data[k];
    if (typeof val === 'object') {
      var td2 = document.createElement('td'); td2.textContent = Number(val.dau).toLocaleString();
      var td3 = document.createElement('td'); td3.textContent = val.community;
      row.appendChild(td1); row.appendChild(td2); row.appendChild(td3);
    } else {
      var td2 = document.createElement('td'); td2.textContent = val + '%';
      row.appendChild(td1); row.appendChild(td2);
    }
    tbody.appendChild(row);
  });
}

function refreshCharts(type) {
  myChart.setOption(buildChart(currentData, currentType));
  myChart2.setOption(buildPenetrationChart(currentPenData, currentType));
}

// 构建标注清单
function buildAnnotationList() {
  var list = document.getElementById('annotationList');
  allAnnotations.forEach(function(a, i) {
    var div = document.createElement('div');
    div.className = 'annotation-item';
    var cb = document.createElement('input');
    cb.type = 'checkbox';
    cb.checked = true;
    cb.addEventListener('change', function() {
      showMap[i] = this.checked;
      refreshCharts();
    });
    var tag = document.createElement('span');
    tag.className = 'tag-' + a.tag;
    tag.textContent = tagLabels[a.tag];
    var date = document.createElement('span');
    date.className = 'ann-date';
    date.textContent = a.date;
    var label = document.createElement('span');
    label.className = 'ann-label';
    label.textContent = a.label;
    div.appendChild(cb);
    div.appendChild(tag);
    div.appendChild(date);
    div.appendChild(label);
    list.appendChild(div);
  });
}

var chartDom = document.getElementById('chart');
var myChart = echarts.init(chartDom);
var chartDom2 = document.getElementById('chart2');
var myChart2 = echarts.init(chartDom2);

var currentData = daysData;
var currentPenData = daysPenetration;
var currentType = 'daily';

myChart.setOption(buildChart(daysData, 'daily'));
myChart2.setOption(buildPenetrationChart(daysPenetration, 'daily'));
renderTable(daysData, 'daily', 'data-table');
renderTable(daysPenetration, 'daily', 'penetration-table');
buildAnnotationList();

document.querySelectorAll('.tab-btn').forEach(function(btn) {
  btn.addEventListener('click', function() {
    document.querySelectorAll('.tab-btn').forEach(function(b) { b.classList.remove('active'); });
    btn.classList.add('active');
    var tab = btn.getAttribute('data-tab');
    if (tab === 'daily') { currentData = daysData; currentPenData = daysPenetration; currentType = 'daily'; }
    else if (tab === 'weekly') { currentData = weeksData; currentPenData = weeksPenetration; currentType = 'weekly'; }
    else { currentData = monthsData; currentPenData = monthsPenetration; currentType = 'monthly'; }

    var thead1 = document.querySelector('#data-table thead tr');
    var thead2 = document.querySelector('#penetration-table thead tr');
    if (tab === 'daily') {
      thead1.innerHTML = '<th>日期</th><th>主动DAU</th><th>社区首页UV</th>';
      thead2.innerHTML = '<th>日期</th><th>功能渗透率</th>';
    } else if (tab === 'weekly') {
      thead1.innerHTML = '<th>周期</th><th>主动DAU(周均)</th><th>社区首页UV(周均)</th>';
      thead2.innerHTML = '<th>周期</th><th>功能渗透率(周均)</th>';
    } else {
      thead1.innerHTML = '<th>月份</th><th>主动DAU(月均)</th><th>社区首页UV(月均)</th>';
      thead2.innerHTML = '<th>月份</th><th>功能渗透率(月均)</th>';
    }

    myChart.setOption(buildChart(currentData, currentType));
    myChart2.setOption(buildPenetrationChart(currentPenData, currentType));
    renderTable(currentData, tab, 'data-table');
    renderTable(currentPenData, tab, 'penetration-table');
  });
});

window.addEventListener('resize', function() { myChart.resize(); myChart2.resize(); });
</script>
</body>
</html>'''

html_path = os.path.join(save_dir, '主动DAU与社区首页UV_折线图.html')
with open(html_path, 'w', encoding='utf-8') as f:
    f.write(html)
print(f'HTML已生成: {html_path}')